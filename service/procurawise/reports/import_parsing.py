"""Fase 23 - pure parsing functions for the requirements import preview: raw
file bytes in, (columns, rows, suggested_mapping) out. No Mongo access here
(same "pure function, no infrastructure" principle as reports/assembly.py)."""

import csv as csv_module
import unicodedata
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook

from procurawise.reports.exceptions import RequirementsImportError

# Target Requirement fields (evaluations.schemas.RequirementCreateRequest)
# mapped to the header synonyms this function recognizes, ES/EN, accent- and
# case-insensitive (via _normalize below). Order matters only for
# readability - matching itself is a dict lookup.
_FIELD_SYNONYMS: dict[str, tuple[str, ...]] = {
    "dimension": ("dimension", "dimension_"),
    "category": ("category", "categoria"),
    "title": ("title", "titulo"),
    "description": ("description", "descripcion"),
    "priority": ("priority", "prioridad"),
    "response_type": ("response_type", "tipo de respuesta", "tipo_respuesta"),
    "weight": ("weight", "peso"),
    "required": ("required", "obligatorio"),
    "buyer_guidance": ("buyer_guidance", "guia", "guia para el comprador"),
}


def _normalize(text: str) -> str:
    stripped_accents = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return stripped_accents.strip().lower()


def _suggest_mapping(columns: list[str]) -> dict[str, str]:
    normalized_columns = {_normalize(c): c for c in columns}
    mapping: dict[str, str] = {}
    for field, synonyms in _FIELD_SYNONYMS.items():
        for synonym in synonyms:
            if synonym in normalized_columns:
                mapping[field] = normalized_columns[synonym]
                break
    return mapping


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv_module.DictReader(StringIO(text))
    columns = list(reader.fieldnames or [])
    rows = [dict(row) for row in reader]
    return columns, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    columns = [str(cell) if cell is not None else "" for cell in header_row]
    rows = []
    for raw_row in rows_iter:
        if all(cell is None for cell in raw_row):
            continue
        rows.append(
            {columns[i]: raw_row[i] for i in range(min(len(columns), len(raw_row))) if columns[i]}
        )
    return columns, rows


def parse_requirements_file(
    filename: str, content: bytes
) -> tuple[list[str], list[dict[str, Any]], dict[str, str]]:
    if len(content) == 0:
        raise RequirementsImportError("el archivo está vacío")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "csv":
        columns, rows = _parse_csv(content)
    elif extension == "xlsx":
        columns, rows = _parse_xlsx(content)
    else:
        raise RequirementsImportError(f"formato no soportado: {extension!r} (usa .xlsx o .csv)")
    if not columns:
        raise RequirementsImportError("el archivo no tiene encabezados")
    return columns, rows, _suggest_mapping(columns)
